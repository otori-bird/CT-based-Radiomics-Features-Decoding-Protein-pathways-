
from PIL import Image
import random
import SimpleITK as sitk
import os
import csv
# import chardet
# import pandas as pd
import numpy as np
# import einops
import openpyxl as op


# new




# feat_path = "../ml_image2pro/total_features.csv"
feat_path =  "../../myFeatureExtractor/post_train_features.csv"
# prot_path = "prot_use.xlsx"
prot_path = "../ml_image2pro/selected_genes.xlsx"
label_path = "../ml_image2pro/group.csv"
psea_path = "../ml_image2pro/data_customized_20.xlsx"
# psea_path = "psea_data.xlsx"


def get_data(feat_path=feat_path,prot_path=prot_path,label_path=label_path):
    # 读取excel文件
    wb = op.load_workbook(prot_path)
    ws = wb.active # 获取当前活动的工作表

    # 创建一个空字典
    pro_dict = {}
    label_dict = {}
    pro_array = []

    # 遍历每一列
    for col in ws.iter_cols():
        # 获取第一行的值作为key
        key = col[0].value
        # 获取其余行的值，并转换成一个数组作为value
        value = np.array([cell.value for cell in col[1:]])
        # 将key和value添加到字典中
        pro_dict[key] = value

    with open(label_path,"r") as f:
        for i, line in enumerate(f.readlines()):
            if i == 0:
                continue
            name, label = line.split(';')
            label_dict[name] = int(label.strip())
    feats = []
    labels = []

    with open(feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                continue
            patient, feat = row[0], row[1:]
            if patient in pro_dict:
                feats.append(np.array(feat,dtype=np.float32))
                pro_array.append(np.array(pro_dict[patient]))
                labels.append(label_dict[patient])
    
    return np.vstack(feats), np.squeeze(np.vstack(pro_array)), np.array(labels)



def get_psea_data(feat_path=feat_path,psea_path=psea_path,label_path=label_path):
    # 读取excel文件
    wb = op.load_workbook(psea_path,read_only=True)
    ws = wb.active # 获取当前活动的工作表

    # 创建一个空字典
    pro_dict = {}
    label_dict = {}
    pro_array = []

    # 遍历每一列
    for row in ws.iter_rows():
        # 获取第一行的值作为key
        key = row[0].value
        # 获取其余行的值，并转换成一个数组作为value
        value = np.array([cell.value for cell in row[1:-1]])
        # 将key和value添加到字典中
        pro_dict[key] = value

    with open(label_path,"r") as f:
        for i, line in enumerate(f.readlines()):
            if i == 0:
                continue
            name, label = line.split(';')
            label_dict[name] = int(label.strip())
    feats = []
    labels = []

    with open(feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                continue
            patient, feat = row[0], row[1:]
            if patient in pro_dict:
                feats.append(np.array(feat,dtype=np.float32))
                pro_array.append(np.array(pro_dict[patient]))
                labels.append(label_dict[patient])
    
    return np.vstack(feats), np.squeeze(np.vstack(pro_array)), np.array(labels)


image_feat_path = "../../myFeatureExtractor/post_test_features.csv"
image_label_path = "../../HCC/2010-2017_label.csv"


key_features_path = "../ml_image2pro/key_features_c0.3.txt"


def get_protein_ct_data(feat_path=feat_path,label_path=label_path,key_feat=False):
    if key_feat:
        with open(key_features_path,"r") as f:
            key_feat_names = [l.strip() for l in f.readlines()]

    label_dict = {}
    with open(label_path,"r") as f:
        for i, line in enumerate(f.readlines()):
            if i == 0:
                continue
            name, label = line.split(';')
            label_dict[name] = int(label.strip())
    feats = []
    key_feats = []
    ex_feats = []
    labels = []
    with open(feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                feat_names = row[1:]
                if key_feat:
                    selected_indicies = [feat_names.index(x) for x in key_feat_names]
            patient, feat = row[0], row[1:]
            if patient in label_dict:
                feats.append(np.array(feat,dtype=np.float32))
                labels.append(label_dict[patient])
                if key_feat:
                    key_feat = [feat[j] for j in range(len(feat)) if j in selected_indicies]
                    ex_feat = [feat[j] for j in range(len(feat)) if j not in selected_indicies]
                    key_feats.append(np.array(key_feat,dtype=np.float32))
                    ex_feats.append(np.array(ex_feat,dtype=np.float32))
    if key_feat:
        return np.vstack(feats), np.vstack(key_feats), np.vstack(ex_feats), np.array(labels)
    else:
        return np.vstack(feats), np.array(labels)
      


def get_image_data(image_feat_path=image_feat_path,image_label_path=image_label_path,key_feat=False):
    # train_path = "../../HCC/2010-2017_train_features.csv"
    # # train_path = "../../myFeatureExtractor/example_train_features.csv"
    # train_label_path = "../../HCC/3_train_label.csv"
    # # train_label_path = "../../HCC/train_label.csv"
    # test_path = "../../HCC/test_features.csv"
    # # test_path = "../../myFeatureExtractor/test_features.csv"
    # test_label_path = "../../HCC/3_test_label.csv"
    if key_feat:
        with open(key_features_path,"r") as f:
            key_feat_names = [l.strip() for l in f.readlines()]

    label_dict = {}
    with open(image_label_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                continue
            patient, label = row[0].split(";")
            label_dict[patient] = int(label)

    feats = []
    labels = []
    key_feats = []
    ex_feats = []
    # patient2val = get_patient2pro_data()
    with open(image_feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                feat_names = row[1:]
                if key_feat:
                    selected_indicies = [feat_names.index(x) for x in key_feat_names]
            patient, feat = row[0], row[1:]
            if patient in label_dict:
                # feat.extend(patient2val[patient])
                feats.append(np.array(feat,dtype=np.float32))
                labels.append(np.array(label_dict[patient]))
                if key_feat:
                    key_feat = [feat[j] for j in range(len(feat)) if j in selected_indicies]
                    ex_feat = [feat[j] for j in range(len(feat)) if j not in selected_indicies]
                    key_feats.append(np.array(key_feat,dtype=np.float32))
                    ex_feats.append(np.array(ex_feat,dtype=np.float32))

    feats = np.vstack(feats)
    labels = np.squeeze(np.vstack(labels))
    if key_feat:
        return feats, np.vstack(key_feats), np.vstack(ex_feats), labels
    else:
        return feats, labels