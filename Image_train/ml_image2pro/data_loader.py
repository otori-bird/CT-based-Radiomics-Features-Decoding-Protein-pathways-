
import csv
# import chardet
# import pandas as pd
import numpy as np
# import einops
import openpyxl as op


# feat_path = "./total_features.csv"
feat_path = "../../myFeatureExtractor/post_train_features.csv"
# prot_path = "prot_use.xlsx"
prot_path = "selected_genes.xlsx"
label_path = "group.csv"
psea_path = "data_customized_top10.xlsx"
# psea_path = "psea_data.xlsx"


def get_data(feat_path=feat_path,prot_path=prot_path,label_path=label_path):
    # 读取excel文件
    wb = op.load_workbook(prot_path)
    ws = wb.active # 获取当前活动的工作表

    # 创建一个空字典
    pro_dict = {}
    label_dict = {}
    pro_array = []

    # 遍历每一列
    for col in ws.iter_cols():
        # 获取第一行的值作为key
        key = col[0].value
        # 获取其余行的值，并转换成一个数组作为value
        value = np.array([cell.value for cell in col[1:]])
        # 将key和value添加到字典中
        pro_dict[key] = value

    with open(label_path,"r") as f:
        for i, line in enumerate(f.readlines()):
            if i == 0:
                continue
            name, label = line.split(';')
            label_dict[name] = int(label.strip())
    feats = []
    labels = []

    with open(feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                continue
            patient, feat = row[0], row[1:]
            if patient in pro_dict:
                feats.append(np.array(feat,dtype=np.float32))
                pro_array.append(np.array(pro_dict[patient]))
                labels.append(label_dict[patient])
    
    return np.vstack(feats), np.squeeze(np.vstack(pro_array)), np.array(labels)



def get_psea_data(feat_path=feat_path,psea_path=psea_path,label_path=label_path):
    # 读取excel文件
    wb = op.load_workbook(psea_path,read_only=True)
    ws = wb.active # 获取当前活动的工作表

    # 创建一个空字典
    pro_dict = {}
    label_dict = {}
    pro_array = []

    # 遍历每一列
    for row in ws.iter_rows():
        # 获取第一行的值作为key
        key = row[0].value
        if key == "":
            key = "pathways"
        # 获取其余行的值，并转换成一个数组作为value
        value = np.array([cell.value for cell in row[1:-1]])
        # 将key和value添加到字典中
        pro_dict[key] = value

    with open(label_path,"r") as f:
        for i, line in enumerate(f.readlines()):
            if i == 0:
                continue
            name, label = line.split(';')
            label_dict[name] = int(label.strip())
    feats = []
    labels = []

    with open(feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        ct_feat_names = []
        for i, row in enumerate(reader):
            if i == 0:
               ct_feat_names = row[1:]
            patient, feat = row[0], row[1:]
            if patient in pro_dict:
                feats.append(np.array(feat,dtype=np.float32))
                pro_array.append(np.array(pro_dict[patient]))
                labels.append(label_dict[patient])
    
    return np.vstack(feats), np.squeeze(np.vstack(pro_array)), np.array(labels), ct_feat_names, pro_dict['pathways'].tolist()


image_feat_path = "../../myFeatureExtractor/post_test_features.csv"
image_label_path = "../../HCC/2010-2017_label.csv"


def get_image_data(image_feat_path=image_feat_path,image_label_path=image_label_path):
    # train_path = "../../HCC/2010-2017_train_features.csv"
    # # train_path = "../../myFeatureExtractor/example_train_features.csv"
    # train_label_path = "../../HCC/3_train_label.csv"
    # # train_label_path = "../../HCC/train_label.csv"
    # test_path = "../../HCC/test_features.csv"
    # # test_path = "../../myFeatureExtractor/test_features.csv"
    # test_label_path = "../../HCC/3_test_label.csv"


    label_dict = {}
    with open(image_label_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                continue
            patient, label = row[0].split(";")
            label_dict[patient] = int(label)

    feats = []
    labels = []
    # patient2val = get_patient2pro_data()
    with open(image_feat_path,"r",encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                continue
            patient, feat = row[0], row[1:]
            if patient in label_dict:
                # feat.extend(patient2val[patient])
                feats.append(np.array(feat,dtype=np.float32))
                labels.append(np.array(label_dict[patient]))

    feats = np.vstack(feats)
    labels = np.squeeze(np.vstack(labels))
    return feats, labels