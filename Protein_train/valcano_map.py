import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from statsmodels.stats.multitest import fdrcorrection
import openpyxl
import os
from collections import defaultdict


# 打开excel文件
# workbook = openpyxl.load_workbook('Kmeans/pathway_analysis.xlsx')
workbook = openpyxl.load_workbook('fdr_pathways.xlsx')

pathways = []
func2category = {}
# 遍历每个工作表，读取第一列数据进行判断
for sheet_name in workbook.sheetnames:
    # if 'Sheet' == sheet_name or '1' in sheet_name:
        # continue
    sheet = workbook[sheet_name]
    p = []
    for cell, cell2 in zip(sheet['A'],sheet['B']):
        if cell.value is not None:
            p.append(cell.value.lower().strip())
            func2category[p[-1]] = [v.strip() for v in cell2.value.lower().strip()[1:-1].split(",")]
    p = p[1:]
    if len(p) > 0:
        pathways.append(p)



root_path = "./Kmeans"
Terms2Pval = defaultdict(list)
Terms2Gene = defaultdict(list)
Terms2FDR = defaultdict(list)

def pathway_condition(item):
    # if float(item['PValue']) > 1e-5:
    #     return False
    if float(item['Bonferroni']) > 0.05:
        return False
    if float(item['FDR']) > 0.05:
        return False
    return True

for subdir, dirs, files in os.walk(root_path):
    for file in files:
        if file.startswith("chart"):
            file_path = os.path.join(subdir, file)
            with open(file_path, "r") as f:
                # 读取第一行作为键值
                keys = f.readline().strip().split('\t')
                
                # 逐行读取，每行以空格拆分并存入字典后添加到列表中
                for line in f:
                    values = line.strip().split('\t')
                    item = {keys[i]: values[i] for i in range(len(keys))}
                    if not pathway_condition(item):
                        continue
                    if '~' in item['Term']:
                        func = item['Term'].split('~')[1].lower()
                    else:
                        func = item['Term'].split(':')[1].lower()
                    func = func.lower().strip()
                    flag = False
                    for p in pathways:
                        if func in p:
                            flag = True
                            break
                    if flag:
                        Terms2Pval[func].append(item['Bonferroni'])
                        Terms2FDR[func].append(item['FDR'])
                        Terms2Gene[func].extend([x.strip() for x in item['Genes'].split(',')])
                

groups = pd.read_csv("../PSEA/PSEA/group_modified.csv",index_col=0)
data = pd.read_excel('../PSEA/PSEA/prot_use_lower.xlsx',index_col=0)

# result_dict = data.to_dict(orient='index')
# for k, v in result_dict.items():
#     result_dict[k] = dict((i, v[i]) for i in data.columns)

# 根据病变等级将病人分组
group_0 = groups[groups['LiverCancerStage'] == 0].index.tolist()
group_1 = groups[groups['LiverCancerStage'] == 1].index.tolist()

threshold = 7
threshold_fc = 0.5850

os.makedirs(f"./Kmeans/fdr_filtered_{threshold}",exist_ok=True)
os.makedirs(f"./Kmeans/fdr_filtered_{threshold}/up",exist_ok=True)
os.makedirs(f"./Kmeans/fdr_filtered_{threshold}/down",exist_ok=True)
os.makedirs(f"./Kmeans/fdr_filtered_{threshold}/pathways",exist_ok=True)

def check(x):
    # print(len([i for i in x if np.abs(i) > 1]))
    return len([i for i in x if np.abs(i) > threshold_fc]) >= threshold

def is_up(x):
    return len([i for i in x if i > threshold_fc]) >= len([i for i in x if i < -threshold_fc])

gene_set = set()
selected_gene = pd.DataFrame()
for i, p in enumerate(pathways):
    for cluster in p:
        try:
            # if cluster == "spliceosome":
            #     print(cluster)
            # 读取数据
            genes_subset = Terms2Gene[cluster]
            # 计算每个基因在两个组中的t-test p-value，并计算fold change
            fc_list = []
            pvalue_list = []

            for g in genes_subset:
                g = g.lower()
                fc = np.log2(data.loc[g, group_1].mean() / data.loc[g, group_0].mean())
                # fc = data.loc[g, group_1].mean() - data.loc[g, group_0].mean()
                ttest = stats.ttest_ind(data.loc[g, group_1], data.loc[g, group_0])
                # stats.ttest_1samp
                pvalue = ttest.pvalue
                fc_list.append(fc)
                pvalue_list.append(pvalue)
            # pvalue_list = fdrcorrection(pvalue_list)[1]

            if not check(fc_list):
                continue
            print(func2category[cluster])

            # 绘制火山图
            fig, ax = plt.subplots(figsize=(10, 8))
            colors = np.where(np.asarray(fc_list) > 0, 'red', 'blue')
            labels = np.where(np.abs(np.asarray(fc_list)) > threshold_fc, genes_subset, '')
            ax.scatter(x=np.asarray(fc_list), y=-1 * np.log10(np.asarray(pvalue_list)), alpha=0.7, s=20, c=np.where(np.abs(np.asarray(fc_list))> threshold_fc, colors, 'grey'))
            for j in range(len(genes_subset)):
                ax.annotate(labels[j], (np.asarray(fc_list)[j], -1 * np.log10(np.asarray(pvalue_list))[j]), fontsize=8, xytext=(-5,5), textcoords='offset points', ha='center', va='bottom')
            ax.axhline(y=-np.log10(0.05), color='grey', linestyle='--')
            ax.axvline(x=threshold_fc, color='grey', linestyle='--')
            ax.axvline(x=-threshold_fc, color='grey', linestyle='--')
            ax.set_xlabel('Fold Change (log2)')
            ax.set_ylabel('-log10 p-value')
            ax.set_title('Volcano Plot')
            file = cluster.replace("/","")
            # plt.savefig(f"./Kmeans/commons/{file}.png")
            # plt.savefig(f"./Kmeans/{i}/{file}.png")
            if is_up(fc_list):
               plt.savefig(f"./Kmeans/fdr_filtered_{threshold}/up/{file}.png")
            else:
               plt.savefig(f"./Kmeans/fdr_filtered_{threshold}/down/{file}.png")
            
            useful_genes = [g  for j, g in enumerate(genes_subset) if np.abs(fc_list[j]) > 1]
            gene_set.update(useful_genes)
            with open(f"./Kmeans/fdr_filtered_{threshold}/pathways/{cluster}.txt","w") as f:
                for g in useful_genes:
                    f.write(g)
                    f.write("\n")
        except Exception as e:
            print(e)
            continue
print(f"Length of total fdr_filtered genes:{len(gene_set)}")
with open(f"./Kmeans/fdr_filtered_{threshold}/genes.txt","w") as f:
    for g in gene_set:
        f.write(g)
        f.write("\n")
        selected_gene = selected_gene.append(data.loc[g,:])
selected_gene.to_excel(f"./Kmeans/fdr_filtered_{threshold}/selected_genes.xlsx")