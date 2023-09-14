# from tsne_torch import TorchTSNE as TSNE
from operator import ge
from re import T
from sklearn.manifold import TSNE
from sklearn.cluster import DBSCAN
import torch
import os
import matplotlib.pyplot as plt
import itertools
import difflib
import numpy as np
import openpyxl
import csv
from tqdm import tqdm
import pickle


pkl_path = "tsne_by_pn.pkl"
protein_number = 4891
patient_number = 68
if os.path.exists(pkl_path):
    with open(pkl_path, 'rb') as f:
        result = pickle.load(f)

    with open("./tsne_plt_data.pkl","rb") as f:
        data = pickle.load(f)
    proteins = data['proteins']
    patients = data['patients']
    labels = data['labels']
    colors = data['colors']
    # labels = {}
    # with open('../PSEA/PSEA/group.csv') as csvfile:
    #     reader = csv.reader(csvfile)
    #     for row in reader:
    #         key, val = row[0].split(';')
    #         if key == "name":
    #             continue
    #         labels[key] = val

    
    # colors = []
    # count = 0
    # for patient, label in labels.items():
    #     if label == '1':
    #         colors.extend(['r'] * protein_number)
    #     else:
    #         colors.extend(['b'] * protein_number)

    # data = {
    #     'proteins':proteins,
    #     'patients':[x for x, _ in labels.items()],
    #     'labels':[x for _, x in labels.items()],
    #     'colors':colors,
    # }

    # with open("./tsne_plt_data.pkl","wb") as f:
    #     pickle.dump(data,f)

else:
    # plt.switch_backend('module://ipympl')
    wb = openpyxl.load_workbook('../protein/prot_clean.xlsx')
    ws = wb.active
    gene2prot = {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        prot, gene = row
        gene2prot[gene] = prot

    wb = openpyxl.load_workbook('../PSEA/PSEA/prot_use.xlsx')
    ws = wb.active
    patient2val = {}
    order = {}
    for i, col in enumerate(ws.iter_cols(min_col=2,values_only=True)):
        patient = col[0]
        patient2val[patient] = col[1:]
        if i == 0:
            for row_idx, cell_value in enumerate(col[1:],start=1):
                gene = ws.cell(row=row_idx+1,column=1).value
                order[gene2prot[gene]] = row_idx-1
            
    labels = {}
    with open('../PSEA/PSEA/group.csv') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            key, val = row[0].split(';')
            labels[key] = val

    data_path = "./merged_emb"
    fasta_embed_files = os.listdir(data_path)
    data = []
    indices = []
    layer = 33 # 0 32 33
    print("Loading Fasta")
    proteins = []
    for i, f in tqdm(enumerate(fasta_embed_files)):
        path = os.path.join(data_path, f)
        if os.path.isfile(path):
            fasta_embed_files[i] = path
        else:
            while not os.path.isfile(path):
                path = os.path.join(path, os.listdir(path)[0])
            fasta_embed_files[i] = path
        prot = path.split('|')[1]
        if prot in order:
            tmp = torch.load(fasta_embed_files[i])
            # prot = tmp['label'].split("|")[1]
            indices.append(order[prot]) 
            proteins.append([order[prot],prot])
            data.append(tmp['mean_representations'][layer])

    proteins = [x for _, x in sorted(proteins)]

    data = torch.vstack(data)
    data_sorted = torch.index_select(data, dim=0, index=torch.tensor(indices))

    print("Combining DATA")

    data = []
    colors = []
    numbers = []
    count = 0
    patients = []
    for patient, label in labels.items():
        if patient in patient2val:
            patient_val = torch.tensor(patient2val[patient])
            data.append(torch.cat([patient_val.unsqueeze(1), data_sorted],dim=1))
            # data.append(data_sorted * patient_val.unsqueeze(1))
            if label == '1':
                colors.extend(['r'] * patient_val.shape[0])
            else:
                colors.extend(['b'] * patient_val.shape[0])
            numbers.extend([str(count)] * patient_val.shape[0])
            patients.append(patient)
            count += 1
    data = torch.vstack(data)

    print("TSNE")
    # tsne = TSNE(n_components=2, perplexity=30, n_iter=1000)
    tsne = TSNE(n_components=2, perplexity=30,)
    result = tsne.fit_transform(data)
    with open(pkl_path,"wb") as f:
        pickle.dump(result,f)


save_path = "./tsne_by_protein"

if not os.path.exists(save_path):
    os.mkdir(save_path)

print("ploting")
# 绘制散点图并添加标签
fig, ax = plt.subplots(figsize=(40,30))

ax.scatter(result[:, 0], result[:, 1],c='gray',s=20)
plt.savefig(os.path.join(save_path,"tsne_by_patient.png"))

objs = []

for i in range(protein_number):
    X_pos = []
    Y_pos = []
    col = []
    for j in range(patient_number):
        X_pos.append(result[j * protein_number + i, 0])
        Y_pos.append(result[j * protein_number + i, 1])
        col.append(colors[j * protein_number + i])
        # objs.append(ax.annotate(proteins[i], (result[j * protein_number + i, 0], result[j * protein_number + i, 1])))
        
    dbscan = DBSCAN(eps=0.5, min_samples=5).fit( np.array( [[x,y] for x, y in zip(X_pos,Y_pos)] ) )
    if len(set(dbscan.labels_)) > 1:
        clusters = dbscan.labels_.tolist()
        min_value = min(clusters)
        max_value = max(clusters)
        classes = max_value - min_value + 1
        groups = [[] for _ in range(classes)]
        level = 0
        for j, c in enumerate(clusters):
            groups[c+min_value].append(j)
        mode_pers = []
        modes = [] # label types
        flag = False
        for k in range(classes):
            l = [labels[j] for j in groups[k]]
            if len(l) > 3:
                mode = max(set(l),key=l.count)
                modes.append(mode)
                mode_count = l.count(mode)
                mode_pers.append(mode_count / len(l))
                if mode == '1' and len(l) > 8:
                    flag = True
        if flag and all(num > 0.5 for num in mode_pers) and len(set(modes)) > 1:
        # if sum(mode_pers) / len(mode_pers) > 0.8 and len(set(modes)) > 1:
            print(mode_pers)
            print(modes)
            print([len(groups[k]) for k in range(classes)])
            print(sum(mode_pers) / len(mode_pers), proteins[i])

            dots = ax.scatter(X_pos, Y_pos,c=col,s=200)
            plt.savefig(os.path.join(save_path,f"{proteins[i]}_tsne_by_patient_with_name.png"))
            dots.remove()

