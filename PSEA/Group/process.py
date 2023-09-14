import openpyxl as op

wb_ref    = op.load_workbook("prot_use.xlsx")
ws_ref    = wb_ref['Sheet1']
wb_source = op.load_workbook("kegg.list.xlsx")

row_m_ref = ws_ref.max_row
col_m_ref = ws_ref.max_column
print("{} rows in reference dataset.".format(row_m_ref))
gene_list = []
for gid in range(0,row_m_ref):
    cell_ref = ws_ref.cell(column = 1, row = gid+1).value
    gene_list.append(cell_ref)

geneid_column = 8
for i in range(0,10):
    curr_sheet = wb_source.sheetnames[i]
    print("Current sheet: {}.".format(curr_sheet))

    ws_curr  = wb_source['{}'.format(curr_sheet)]
    row_m    = ws_curr.max_row
    column_m = ws_curr.max_column
    print("sheet max column {}.".format(column_m))
    print("sheet max row {}.".format(row_m))

    hub = []
    for j in range(0,row_m):
        if j == 0:
            pass
        else:
            item = ws_curr.cell(column = geneid_column, row = j+1).value
            gene_id = item.split('/')

            for k in gene_id:
                if hub.count(k) > 0:
                    print ("{} exsits!".format(k))
                    pass
                else:
                    print("{} doesn't exsit, push in.".format(k))
                    hub.append(k)
            
    print("Gene id list is {}.".format(hub))
    print("Try to find gene from reference.")
    index_list = []
    gene_find = 0
    gene_miss = 0
    for target in hub:
        if gene_list.count(target) > 0:
            pos = gene_list.index(target)
            index_list.append(pos)
            gene_find = gene_find+1
        else:
            gene_miss = gene_miss+1

    print("{} genes were find among {} targets.".format(gene_find,len(hub)))
    print("{} genes missed.".format(gene_miss))
    print(index_list)
    
    print("write GP file.")
    wbs = op.Workbook()
    write_file = wbs.active

    col_wf = 1
    for row_index_ref in index_list:
        row_index_ref = int(row_index_ref)+1
        row_wf = 1
        for col_index_ref in range(0,col_m_ref):
            data = ws_ref.cell(column = col_index_ref+1, row = row_index_ref).value
            print(data)
            write_file.cell(row=row_wf, column=col_wf, value=data)
            row_wf = row_wf+1
        col_wf = col_wf+1
        
    print("save xlsx.")
    wbs.save("{}.xlsx".format(wb_source.sheetnames[i]))





            

